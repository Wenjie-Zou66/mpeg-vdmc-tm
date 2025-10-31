
#!/usr/bin/env python3
import os
import sys
import argparse
import zipfile

def getParentDir():
  return os.path.dirname(os.path.dirname( os.path.realpath(__file__) ) )

def getCurrentDir():
  return os.path.dirname( os.path.realpath(__file__) )

# Import openpyxl module
try:
  sys.path.append( getCurrentDir() + "/.modules/" )
  from openpyxl import load_workbook
  from openpyxl.utils import get_column_letter
except ImportError as error:
  print("Install openpyxl:")
  os.system("pip3 install openpyxl --upgrade --target=%s \n" % ( getCurrentDir() + "/.modules/" ) )
  print("openpyxl installed, please restart script")
  exit()

# Set configurations
seqOrder = [0, 1, 2, 3, 6, 7, 4, 5, 8]
maxFrames = [300, 300, 300, 300, 300, 150, 300, 300, 1]
strCond = [{'name': 'C0 lossless AI', 'dataColumn': [10, 17], 'frameColumn': 5},
           {'name': 'C1 lossy AI',    'dataColumn': [14, 35], 'frameColumn': 6},
           {'name': 'C2 lossy RA',    'dataColumn': [14, 35], 'frameColumn': 6}]
losslessColumn = ['TotalBitstreamBits', 'BaseMeshIntraBits', 'UserEncoderRuntime', 'UserDecoderRuntime']
lossyColumn = ['NbOutputFaces', 'TotalBitstreamBits', 'BaseMeshIntraBits', 'BaseMeshInterBits', 'DisplacementBits', \
              'AttrVideoBits','MetaDataBits','GridD1', 'GridD2', 'GridLuma', 'GridChromaCb', 'GridChromaCr', \
              'IbsmGeom', 'IbsmLuma', 'UserEncoderRuntime', 'UserDecoderRuntime']

def parseArgs():
  global parser
  source = getParentDir() + "/results/empty.xlsm"
  parser = argparse.ArgumentParser(description="Fill CTC .xlsm file")
  # clang-format off
  parser.add_argument("--anchor", help="Anchor .csv path (csv)",            default="anchor.cvs", type=str)
  parser.add_argument("--tested", help="Test.csv path (csv)",               default="test.csv",   type=str)
  parser.add_argument("--source", help="VMesh CFP spreadsheet path (xlsm)", default=source,       type=str)
  parser.add_argument("--frame",  help="Number of frames",                  default=-1,           type=int)
  parser.add_argument("--save",   help="Create spreadsheet path (xlsm)",    default="test.xlsm",  type=str)
  # clang-format one
  return parser.parse_args()

def printArgs(args):
  print("Argument values:")
  for arg in vars(args):
    print("  - %-20s = %s" % (arg, getattr(args, arg)))

# DataFrame to read our input CS file
def readCsv(filename):
  from csv import DictReader
  with open( filename, 'r') as read_obj:
    data = []
    for row in DictReader(read_obj):
      data.append( row )
  return data

# Print sequences
def printData(data):
  print("data:", data)
  for line in data:
    print(" - Cond = %s seq = %s rate = %s: %s " %
      (line["CondId"], line["SeqId"], line["RateId"], line))

# Get results of one test
def get(array, condId, seqId, rateId):
  for el in array:
    if (int(el['CondId']) == condId and int(el['SeqId']) == seqId and int(el['RateId']) == rateId):
      return (el)
  return ()

# Update frame number in xlsm output file
def updateFrameNumber(wb, frame):
  for condId in range(0, 3):
    ws = wb[strCond[condId]['name']]
    numRate =  1 if (condId == 0) else 5
    for idx, seqId in enumerate(seqOrder):
      for rate in range(0, numRate):
        ws.cell( 5 + idx * numRate + rate, strCond[condId]['frameColumn']).value = int( min( frame, maxFrames[idx] ) )

#  Update the results of one experiment
def update(wb, path, testId=0):
  data = readCsv(path)
  ws = wb['Summary']
  ws.cell(1 + testId, 3).value = os.path.splitext(os.path.basename(path))[0]
  for condId in range(0, 3):
    ws = wb[strCond[condId]['name']]
    columnId = strCond[condId]['dataColumn'][testId]
    for idx, seqId in enumerate(seqOrder):
      if (condId == 0):
        el = get(data, condId, seqId + 1, 0)
        if (len(el) > 0):
          for i in range(0, len(losslessColumn)):
            cell = ws.cell(5 + idx, columnId + i)
            value = float( el[losslessColumn[i]] )
            cell.value = 'ERROR' if value<0 else value if i>1 else int(value)
      else:
        for rateId in range(0, 5):
          el = get(data, condId, seqId + 1, rateId + 1)
          if (len(el) > 0):
            for i in range(0, len(lossyColumn)):
              cell = ws.cell(5 + idx * 5 + rateId, columnId + i)
              value = float(el[lossyColumn[i]])
              cell.value = 'ERROR' if value<0 else value if i>1 else int(value)

# Fill xlsm file
def fillXlsm(source, output, anchor, tested, frame):
  wb = load_workbook(source, read_only=False, keep_vba=True)
  updateFrameNumber(wb, frame)
  update(wb, anchor, 0)
  update(wb, tested, 1)
  wb.save(output)

if __name__ == "__main__":
  # Parse arguments
  args = parseArgs()

  # Check options
  if args.frame <= 0:
    printArgs(args)
    print("Error: number of frame must be set \n" )
    exit(-1)
  if not os.path.isfile( args.source ) and not os.path.exists( args.source ):
    printArgs(args)
    print("Error: source xlsm file not exist: %s \n"% args.source)
    exit(-1)
  if not os.path.isfile( args.anchor ) and not os.path.exists( args.anchor ):
    printArgs(args)
    print("Error: anchor csv file not exist: %s \n"% args.anchor)
    exit(-1)

  # Fill xlsm
  fillXlsm( args.source, args.save, args.anchor, args.tested, args.frame )
  print("%s create %s " % ( os.path.basename(__file__), args.save) )
